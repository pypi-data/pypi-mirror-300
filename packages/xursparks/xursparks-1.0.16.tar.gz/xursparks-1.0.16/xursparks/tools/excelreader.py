import os
from dataclasses import dataclass
from itertools import islice

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.worksheet.cell_range import CellRange

import pandas as pd

from typing import Dict, List, Any, Tuple

@dataclass
class Keyword:
    keyword: str
    col: str = None
    row: str = None
    col_property: str = None
    row_property: str = None


class FileReader:
    """Class that helps extracts tables from an excel file"""
    def __init__(self, 
                 filename: str, 
                 sheetname: List[str]|str = None,
                 column_num : int = None,
                 row_num: int = None,
                 table_start_column: str = None,
                 table_end_column: str = None,
                 table_start_row: str = None,
                 table_end_row: str = None,
                 keyword_start: str = None, 
                 keyword_end: str = None, 
                 keyword_start_column: str = None,
                 keyword_end_column: str = None,
                 keyword_start_row: str|int = None,
                 keyword_end_row: str|int = None,
                 keyword_start_property: str | Dict[str, str] = None,
                 keyword_end_property: str | Dict[str, str] = None,
                 pandas_parameters: Dict[str, Any] = None,
                 tolerance: int = 0,
                 unmerge_and_fill: bool = True,
                 **kwargs):
        """Initializes a FileReader Class.
        Input:
            filename: Path of the excel file
            sheetname: Name of the sheet/s to extract data from. Can be a list
            column_num: Number of Columns. 
            row_num: Number of Rows.
            table_start_column: Column in which the table starts.
                                If None, it is assumed to be the same column as with keyword_start.
            table_start_end: Row in which the table starts.
                             If None, it is assumed to be the same row as with keyword_start.
            
            keyword_start: Keyword (if any) to be used to find the start of the table
            keyword_end: Keyword (if any) to be used to find the end of the table
            keyword_start_column: Column Name/Header (if any) to be used to find start keyword.
                                - If None, searches whole sheet for the keyword.
            keyword_end_column: Column Name/Header (if any) to be used to find end keyword.
                                - If None, searches all available columns for the keyword.
            keyword_start_row: Row Number/Name (if any) to be used to find start keyword.
                                - If None, searches all available rows for the keyword.
                                            - If None, searches all available columns for the keyword.
            keyword_end_row: Row Number/Name (if any) to be used to find end keyword.
                                - If None, searches all available rows for the keyword.
            keyword_start_property: Accepts either None, 'after', 'before', and dict with keys 'row', and 'col'.
                                - If None, keyword indicates the start of the data table.
                                - If 'before', keyword indicates the start of the data table is row before keyword.
                                - If 'after', keyword indicates the start of the table is row after keyword.
                                - If dict is inputted, keys should be 'row' and 'col'. This is so that the parameters could
                                  be more specific on whether the before or after keyword means before or after the row or column of keyword.
            keyword_end_property: The same as keyword_start property except indicates the end of the table.
            tolerance: tolerance for blank rows/columns when searching for table
            unmerge_and_fill: If True, unmerges and fills all merged cells. Defaults to True
        """
        self.filename = filename
        self.sheetname = sheetname
        self.tolerance = tolerance
        
        self.table_dimensions = {'num_col': column_num,
                                 'num_row': row_num,
                                 'start_col': table_start_column,
                                 'start_row': table_start_row,
                                 'end_col': table_end_column,
                                 'end_row': table_end_row}

        if not isinstance(keyword_start_property, dict) and keyword_start_property not in [None, 'before', 'after']:
            raise ValueError("Invalid value for 'keyword_start_property'. Should be None, 'before' or 'after")
        elif isinstance(keyword_start_property, dict):
            for key, value in keyword_start_property.items():
                if key not in ['row', 'col']:
                    raise ValueError("Invalid value for keys in 'keyword_start_property'. Should be, 'row', 'col'")
                elif value not in [None, 'before', 'after']:
                    raise ValueError(f"Invalid value for 'keyword_start_property['{key}']'. Should be None, 'before' or 'after")
        
        if not isinstance(keyword_end_property, dict) and keyword_end_property not in [None, 'before', 'after','all']:
            raise ValueError("Invalid value for 'keyword_end_property'. Should be None, 'before' or 'after")
        elif isinstance(keyword_end_property, dict):
            for key, value in keyword_end_property.items():
                if key not in ['row', 'col']:
                    raise ValueError("Invalid value for keys in 'keyword_end_property'. Should be, 'row', 'col'")
                elif value not in [None, 'before', 'after', 'all']:
                    raise ValueError(f"Invalid value for 'keyword_end_property['{key}']'. Should be None, 'before' or 'after")
        
        def initialize_keyword(keyword, col, row, params):
            word = Keyword(keyword=keyword,
                          col=col,
                          row=row)
            if params is not None:
                if isinstance(params, dict):
                    for key, value in params.items():
                        if key =='col':
                            word.col_property = value
                        elif key == 'row':
                            word.row_property = value
                else:
                    word.row_property = params


            return word
        self.keyword = dict()
        self.keyword['start'] = initialize_keyword(keyword_start,
                                                    keyword_start_column,
                                                    keyword_start_row,
                                                    keyword_start_property)
            
        self.keyword['end'] = initialize_keyword(keyword_end,
                                                  keyword_end_column,
                                                  keyword_end_row,
                                                  keyword_end_property)
        
        self.pandas_parameters = {} if pandas_parameters is None else pandas_parameters
        self.unmerge_and_fill = unmerge_and_fill
        self.__load_file__()

    def __load_file__(self):
        self.wb = openpyxl.load_workbook(self.filename, data_only=True)
        self.sheet = self.wb[self.sheetname]

    def __find_loc__(self, position:str,property:str, cell:openpyxl.cell.cell.Cell, dimension: str):
        if dimension == 'col':
            if self.table_dimensions[f'{position}_col'] is None:
                if property == 'before':
                    loc = get_column_letter(cell.offset(column=-1).column)
                elif property == 'after':
                    loc = get_column_letter(cell.offset(column=+1).column)
                elif property is None:
                    loc = get_column_letter(cell.column)
            else:
                loc = self.table_dimensions[f'{position}_col']

        elif dimension =='row':
            if self.table_dimensions[f'{position}_row'] is None:
                if property == 'before':
                    loc = cell.row - 1
                elif property == 'after':
                    loc = cell.row + 1
                else:
                    loc = cell.row
            else:
                loc = self.table_dimensions['start_row']

        else:
            raise ValueError(f"'dimension' must be either 'row', or 'col'. Not {dimension}")
        return loc
    
    def __find_loc_all__(self, cell:openpyxl.cell.cell.Cell, start_location: openpyxl.cell.cell.Cell, dimension:str, opp_param:str=None, start:bool=False, params:dict={})->openpyxl.cell.cell.Cell:
        """if start is True, then tries to find the end of the table without any end condition, except tolerance starting from start location"""
        print(f"start location: {start_location}")
        print(f"start loc new: {column_index_from_string(start_location[0])}")
        params['min_col'] = column_index_from_string(start_location[0]),
        params['min_row'] = start_location[1]

        if not start:
            if dimension == 'col':
                key = 'max_row'
            elif dimension == 'row':
                key = 'max_col'

            if key not in params:
                if opp_param == 'before':
                    params[key] = cell.row - 1
                elif opp_param is None:
                    params[key] = cell.row
                elif opp_param == 'after':
                    params[key] = cell.row + 1

        current_cell = None
        iterator = self.sheet.iter_cols if dimension == 'col' else self.sheet.iter_rows
        consecutive_blank_count = 0
        print(f"params: {params}")
        for row in iterator(**params):
            if all(cell.value is None for cell in row):
                consecutive_blank_count += 1
                if consecutive_blank_count > self.tolerance:
                    break
            else:
                consecutive_blank_count = 0
                current_cell = row[-1]


        return get_column_letter(current_cell.column), current_cell.row

    def __get_location__(self, position:str, cell:openpyxl.cell.cell.Cell, start_location:openpyxl.cell.cell.Cell=None)->openpyxl.cell.cell.Cell:
        ATTRIBUTES = ['col', 'row']
        for key in ATTRIBUTES:
            value = getattr(self.keyword[position], f"{key}_property")
            if position == 'end' and value == 'all':
                col, row = self.__find_loc_all__(cell, start_location,key, self.keyword[position].col_property if key == 'row' else self.keyword[position].row_property)

            elif key == 'col':
                col = self.__find_loc__(position, value, cell,key)
            elif key == 'row':
                row = self.__find_loc__(position, value, cell, key )

        cell_pos = str(col)+str(row)

        return self.sheet[cell_pos]

    def __find_table_position__(self, 
                                position:str, 
                                start_location:tuple=None, 
                                start_row:str=None, 
                                start_col:str=None, 
                                num_row:int=None, 
                                num_col:int=None, 
                                end_row:str=None, 
                                end_col:str=None, 
                                **kwargs)->tuple:
        
        """Finds start or end location of table."""
        def iterate_cells(keyword:Any, sheet: openpyxl.worksheet.worksheet.Worksheet, dimension:str, col:str=None,params:dict=None):
            params = dict() if params is None else params
            if position == 'end':
                params['min_col'] = column_index_from_string(start_location[0])
                params['min_row'] = start_location[1]

            if dimension == 'rows':
                for row in sheet.iter_rows(**params):
                    for cell in row:
                        if cell.value == keyword:
                            return cell
                        
            elif dimension == 'cols' and col is not None:
                for cell in sheet[col]:
                    if cell.value == keyword:
                        return cell
                    
            elif dimension == 'cols':
                for col in sheet.iter_cols(**params):
                    for cell in col:
                        if cell.value == keyword:
                            return cell 
            else:
                return None

        if position not in ['start', 'end']:
            return ValueError(f"'position' parameter should either be 'start' or 'end'. Not {position}")

        if position == 'start' and start_col is not None and start_row is not None:
            return (start_col, start_row)
        
        elif position == 'end':
            if end_col is not None and start_row is not None:
                return (end_col, end_row)

            if num_col is not None and num_row is not None:
                return (get_column_letter(column_index_from_string(start_location[0])+num_col), start_location[1]+num_row)
            
            if self.keyword['end'] is None:
                """        end_location = self.__find_loc_all__(self.sheet[start_location[0]+str(start_location[1])],
                                                 start_location=start_location,
                                                 dimension='col',
                                                 start=True)"""
                params = {}
                if num_row is not None:
                    params['max_row'] == start_location[1] + num_row -1
                if num_col is not None:
                    params['max_col'] == start_location[0] + num_col -1
                if end_row is not None:
                    params['max_row'] == end_row
                if end_col is not None:
                    params['max_col'] == column_index_from_string(end_col)


                self.__find_loc_all__(self.sheet[start_location[0]+str(start_location[1])],
                                      start_location=start_location,
                                      params=params)

        position_cell = None
        if self.keyword[position].col is not None:
            if self.keyword[position].row is not None:
                if position=='end':
                    params={
                        'min_row': start_location[1]+1,
                        'min_col': self.sheet[self.keyword[position].col][0].column,
                        'max_col': self.sheet[self.keyword[position].col][0].column if end_col is None else end_col
                    }
                    position_cell = iterate_cells(keyword=self.keyword[position].keyword,sheet= self.sheet, dimension='cols', params=params)
                else:
                    position_cell = iterate_cells(keyword=self.keyword[position].keyword,sheet= self.sheet, dimension='cols', col= self.keyword[position].col)
                
                if position_cell is None:
                    raise ValueError(f"Keyword {self.keyword[position].keyword} not found in {self.keyword[position].col}.")
                else:
                    cellname = str(self.keyword[position].col) + str(self.keyword[position].row)
                    position_cell = self.sheet[cellname]
            else: #if keyword[position].row is None
                position_cell = iterate_cells(keyword=self.keyword[position].keyword,
                                              sheet=self.sheet,
                                              dimension='cols',
                                              col=self.keyword[position].col)

        elif self.keyword[position].row is not None:
            params = {
                'min_row': int(self.keyword[position].row),
                'max_row': int(self.keyword[position].row)
            }
            position_cell = iterate_cells(keyword=self.keyword[position].keyword, sheet=self.sheet, dimension='cols', params=params)

        else: #iterates through all cells if no keyword location is specified
            position_cell = iterate_cells(keyword=self.keyword[position].keyword, sheet=self.sheet, dimension='cols')


        location = self.__get_location__(position, position_cell, start_location)

        if position == 'start':
            if start_col is not None:
                return start_col, location.row
            elif start_row is not None:
                return (get_column_letter(location.column), start_row)
        elif position == 'end':    
            if end_col is not None:
                return (end_col, location.row)
            elif end_row is not None:
                return (get_column_letter(location.col), end_row)
            elif num_col is not None:
                return (get_column_letter(start_location[0]+num_col-1), location.row)
            elif num_row is not None:
                return (get_column_letter(location.col), start_location[1]+num_row-1)
        
        return (get_column_letter(location.column), location.row)

    def __find_table_range__(self) -> Tuple[Tuple[int,int], Tuple[int,int]]:
        
        start_location = self.__find_table_position__('start', **self.table_dimensions)

        end_location = self.__find_table_position__('end', start_location = start_location, **self.table_dimensions)

            
        return start_location, end_location

    def __get_merged_cell_ranges__(self)->List:
        merged_cells = []
        for merged_cell in self.sheet.merged_cells.ranges:
            merged_cells.append(merged_cell)

        return merged_cells
        
    def __get_merged_cell_ranges_in_table_range__(self, merged_cells_range:List[CellRange], table_range:str):
        table_range = CellRange(table_range)
        filtered_merged_cells = []
        for merged_cell in merged_cells_range:
            if table_range.issuperset(merged_cell):
                filtered_merged_cells.append(merged_cell)

        return filtered_merged_cells
    
    def unmerge_and_fill_cell(self,cell_range:str):
        # Get the top-left cell of the merged range
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        top_left_cell = self.sheet.cell(min_row, min_col)
        value = top_left_cell.value

        # Unmerge the cells
        self.sheet.unmerge_cells(cell_range)

        # Fill all cells in the previously merged range with the value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                self.sheet.cell(row, col).value = value

    def get_table(self) -> pd.DataFrame:
        """Finds the data table based on parameters given."""
        if self.keyword['start'].keyword is None and self.keyword['end'].keyword is None:
            ### ingests whole sheet
            if self.unmerge_and_fill:
                ranges = self.__get_merged_cell_ranges__()
                for cell_range in ranges:
                    self.unmerge_and_fill_cell(cell_range.coord)
            
            data = self.sheet.values
            cols = next(data)[1:]
            data = list(data)
            idx = [r[0] for r in data]
            data = (islice(r, 1, None) for r in data)
            df = pd.DataFrame(data, index=idx, columns=cols, **self.pandas_parameters)
        else:
            ### ingests only range
            start_location, end_location = self.__find_table_range__()
            table_range = f"{start_location[0]}{start_location[1]}:{end_location[0]}{end_location[1]}"

            if self.unmerge_and_fill:
                ranges = self.__get_merged_cell_ranges__()
                ranges = self.__get_merged_cell_ranges_in_table_range__(ranges, table_range)
                for cell_range in ranges:
                    self.unmerge_and_fill_cell(cell_range.coord)

            cell_range = self.sheet[table_range]
            values_list = []
            for row in cell_range:
                row_values = [cell.value for cell in row]
                values_list.append(row_values)
            df = pd.DataFrame(values_list[1:], columns=values_list[0], **self.pandas_parameters)
        
        return df
    
    def get_keywords(self, keyword:str, row:str=None, col:str=None, first_instance_only:bool=True, exact_match:bool=False)->List[Keyword]:
        """"Gets keywords if there are multiple in one sheet. Returns a list of Keyword class instances.
        Checks if the keyword 'string' is found in the cell value.
        Example: 
            Keyword: "BCD"
            Will return: "ABCDEF"
        Input:
            keyword: The keyword string to be found
            row: The row where the keyword could be found
            first_instance_only: boolean. Returns the first exact instance of a keyword if True, else returns all instances.
        """
        keywords = []

        params = {}
        if row is not None:
            params['min_row'] = row
            params['max_row'] = row
        
        if col is not None:
            params['min_col'] = column_index_from_string(col)
            params['max_col'] = column_index_from_string(col)

        if first_instance_only:
            existing_keywords = set()

        for row in self.sheet.iter_cols(**params):
            for cell in row:
                cell_value = str(cell.value)
                if keyword is not None and keyword in cell_value:
                    if not first_instance_only or (first_instance_only and cell.value not in existing_keywords):
                        keywords.append(Keyword(
                            keyword=cell.value,
                            col=cell.column_letter,
                            row=cell.row
                        ))
                        if first_instance_only:
                            existing_keywords.add(cell.value)
                    
        if len(keywords) == 0:
            if row is not None:
                if col is not None:
                    print(f"No Keywords found in {col}{row}")
                else:
                    print(f"No Keywords found in Row {row}")
            else:
                print(f"No Keyword found in Column {col}")     

        return keywords