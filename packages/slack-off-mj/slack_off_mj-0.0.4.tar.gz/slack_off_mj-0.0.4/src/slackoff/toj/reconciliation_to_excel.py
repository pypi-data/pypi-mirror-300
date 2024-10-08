from abc import ABC, abstractmethod
from openpyxl import load_workbook
from os.path import join
from re import match


class ToWriter(ABC):
    @abstractmethod
    def write(self, data_source):
        pass


# 写入到Excel
class ToExcel(ToWriter):
    def __init__(self, rules_):
        if isinstance(rules_, dict):
            self.rules = [rules_]
        elif isinstance(rules_, list):
            self.rules = rules_
        else:
            raise TypeError()

    @staticmethod
    def load_xlsx(rules):
        target = rules["target"]
        if "dir" not in target:
            raise ValueError("target must contain 'dir'")
        if "filename" not in target:
            raise ValueError("target must contain 'filename'")
        xlsx_path = join(target["dir"], target["filename"])
        if not isinstance(xlsx_path, str):
            raise ValueError("target must contain 'filename'")
        if "sheet" not in target:
            raise ValueError("target must contain 'sheet'")
        sheet_params = target["sheet"]
        if "name" not in sheet_params:
            raise ValueError("target must contain 'name'")
        workbook = load_workbook(str(xlsx_path))
        return [workbook[sheet_params["name"]], workbook, xlsx_path]

    def write(self, data_source):
        for rule in self.rules:
            if "target" not in rule:
                raise ValueError("rules_ must contain 'target'")
            if "writer" not in rule:
                raise ValueError("rules_ must contain 'writer'")
            if not isinstance(data_source, dict):
                raise ValueError()
            if "writer" not in rule:
                raise ValueError("rules must contain 'writer'")
            writer_rules = rule["writer"]
            if not (isinstance(writer_rules, dict) or isinstance(writer_rules, list)):
                raise ValueError()
            if isinstance(writer_rules, dict):
                writer_rules["writer"] = [writer_rules["writer"]]
            excel = self.load_xlsx(rule)
            for writer_rule in writer_rules:
                self.doWrite(writer_rule, data_source, excel[0])
            workbook = excel[1]
            workbook.save(excel[2])
            workbook.close()

    def doWrite(self, writer_rule, data_source, sheet):
        if "row" not in writer_rule:
            raise ValueError("rules must contain 'row'")
        row_chain = writer_rule["row"]
        if not isinstance(row_chain, str):
            raise ValueError("row must contain 'str'")
        if "col" not in writer_rule:
            raise ValueError("rules must contain 'col'")
        col = writer_rule["col"]
        if not isinstance(col, dict):
            raise ValueError("col must contain 'dict'")
        if not all(match(r"\d+", c) for c in col.keys()):
            raise ValueError("col must contain 'int'")
        start_row_num = 0
        if "start_row_num" in writer_rule:
            start_row_num = writer_rule["start_row_num"]

        row_num = 1
        row_chain = row_chain.split(".")
        for vk, ve in data_source.items():  # 数据
            row_data = {vk: ve}
            col_keys = col.keys()
            if len(row_chain) > 0:
                row_data = self.to_(row_chain, row_data)  # 确定数据区间
            if row_data is None:
                continue

            for col_key in col_keys:  # 获取列数据
                col_exper = col[col_key]
                col_value = None
                if callable(col_exper):
                    col_value = col_exper(row_num + start_row_num, row_data)
                elif isinstance(col_exper, str):
                    col_value = self.to_(col_exper.split("."), row_data)
                elif not (callable(col_exper) or isinstance(col_exper, str)):
                    raise ValueError("列规则类型只能是字符串和函数类型")
                sheet.cell(row=row_num + start_row_num, column=int(col_key), value=col_value)
            row_num += 1

    @staticmethod
    def to_(row_chain, obj):
        if isinstance(obj, dict):
            row_data = obj.copy()
            # 确定写入行区间
            each_key_stack = [item for item in row_data.keys()]
            row_chain_i = 0
            for target_key in row_chain:
                temp_key_stack = []
                for source_key in each_key_stack:
                    if match(target_key, source_key) is not None:
                        if isinstance(row_data[source_key], dict):
                            temp_key_stack.append(row_data[source_key].keys())
                        row_data = row_data[source_key]
                        row_chain_i += 1
                        break
                each_key_stack.clear()
                each_key_stack = temp_key_stack

            if row_chain_i < len(row_chain):
                return None
            return row_data
        return None
