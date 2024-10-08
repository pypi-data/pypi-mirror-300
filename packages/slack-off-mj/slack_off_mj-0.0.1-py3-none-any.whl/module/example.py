from src.module.reconciliation import XlsxReconciliation
from openpyxl import load_workbook
from re import findall


def init_xlsx_iter(ctx):
    if ctx.current_file_index >= len(ctx.files):
        ctx.iter = None
    else:
        filepath = ctx.files[ctx.current_file_index]
        ctx.current_file_index += 1
        workbook = load_workbook(filename=filepath)
        sheet = workbook["Sheet"]
        ctx.iter = sheet.iter_rows(min_row=2)


def jymx_file(ctx):
    row = ctx["table"]
    if row[18] is None or row[24] is None:
        return True
    if "银联商务" in row[17]:
        return False
    if "财付通" in row[18] and "1656660632" in row[24]:
        return False
    return True


def jymx_kaihu(ctx):
    print(ctx["table"][13])
    return "CBS"


def actual_picture(ctx):
    picture = ctx["table"][13]
    return picture


xlsx = XlsxReconciliation({
    "generator": [
        {
            "source": {
                "init_iterator": init_xlsx_iter,
                "iterator_item": lambda cells: [cell.value for cell in cells],
                "filter": lambda filename: filename.endswith(".xlsx") and "泊时易" in filename,
                "folders": ["E:\\YLSW\\工作项目\\交接的项目\\北投富能RPA流程\\对账结果\\"]
            },
            "rules": {
                "rule": lambda ctx: findall(r"\d{4}\d{2}\d{2}", ctx["filename"])[0],
                "children": [
                    {
                        "rule": lambda ctx: ctx["table"][1],
                        "children": [
                            {
                                "rule": lambda ctx: ctx["table"][2]
                            }
                        ]
                    }
                ]
            }
        },
        {
            "source": {
                "init_iterator": init_xlsx_iter,
                "iterator_item": lambda cells: [cell.value for cell in cells],
                "filter": lambda filename: "交易明细清单.xlsx" in filename,
                "folders": ["E:\\YLSW\\工作项目\\交接的项目\\北投富能RPA流程\\对账结果\\"]
            },
            "rules": {
                "filter": jymx_file,
                "rule": lambda ctx: findall(r"\d{4}-\d{2}-\d{2}", ctx["table"][2])[0].replace("-", ""),
                "children": [
                    {
                        "rule": "actual",
                        "children": [
                            {
                                "rule": actual_picture
                            }
                        ]
                    }
                ]
            }
        }
    ],
    "merge": {
        "\\d{8}": {
            "exper": "${泊时易} * 0.05",
            "name": "管理费"
        }
    },
    "toXlsx": [
        {
            "target": {
                "sheet": {
                    "name": "Sheet1"
                },
                "dir": "E:\\YLSW\\工作项目\\交接的项目\\北投富能RPA流程\\对账结果",
                "filename": "夏娃.xlsx"
            },
            "writer": [
                {
                    "start_row_num": 1,
                    "row": "\\d{8}",
                    "col": {
                        "1": "泊时易",
                        "2": lambda num, row: "=SUM(A" + str(num) + ")"
                    }
                }
            ]
        }
    ]
})

print(xlsx.handle())
