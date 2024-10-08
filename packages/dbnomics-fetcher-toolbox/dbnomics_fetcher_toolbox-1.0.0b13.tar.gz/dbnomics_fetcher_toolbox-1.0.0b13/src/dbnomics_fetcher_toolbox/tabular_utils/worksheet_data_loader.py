from collections.abc import Iterator
from dataclasses import dataclass
from typing import TYPE_CHECKING

from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from .errors import ExpectedEmptyCell, UnexpectedCellValue

if TYPE_CHECKING:
    from openpyxl.cell import Cell


__all__ = ["WorksheetDataLoader"]


@dataclass(frozen=True)
class WorksheetDataLoader:
    worksheet: Worksheet

    def check_cell_text(self, coordinates: str, text: str) -> None:
        ws = self.worksheet

        cell = ws[coordinates]
        if isinstance(cell, tuple):
            msg = f"Expected cell coordinates, got {coordinates!r}"
            raise TypeError(msg)

        value = cell.value
        if not isinstance(value, str) or value != text:
            raise UnexpectedCellValue(cell=cell, expected_cell_value=text, worksheet=ws)

    def check_empty_cell(self, coordinates: str) -> None:
        ws = self.worksheet

        cell = ws[coordinates]
        if cell.value is not None:
            raise ExpectedEmptyCell(cell=cell, worksheet=ws)

    def check_empty_column(self, column: str, *, from_row: int | None = None, to_row: int | None = None) -> None:
        ws = self.worksheet

        column_index = column_index_from_string(column)
        for (cell,) in ws.iter_rows(min_row=from_row, max_row=to_row, min_col=column_index, max_col=column_index):
            if cell.value is not None:
                raise ExpectedEmptyCell(cell=cell, worksheet=ws)

    def check_empty_row(self, row: int, *, from_column: str | None = None, to_column: str | None = None) -> None:
        ws = self.worksheet

        from_column_index = None if from_column is None else column_index_from_string(from_column)
        to_column_index = None if to_column is None else column_index_from_string(to_column)
        for (cell,) in ws.iter_cols(min_col=from_column_index, max_col=to_column_index, min_row=row, max_row=row):
            if cell.value is not None:
                raise ExpectedEmptyCell(cell=cell, worksheet=ws)

    def iter_non_empty_cells(self, row: int, *, from_column: str | None = None) -> Iterator["Cell"]:
        ws = self.worksheet

        from_column_index = None if from_column is None else column_index_from_string(from_column)
        for (cell,) in ws.iter_cols(min_col=from_column_index, min_row=row, max_row=row):
            if cell.value is None:
                break

            yield cell
