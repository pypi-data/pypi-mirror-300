import pandas as pd
import os
from openpyxl import load_workbook

def log_results_to_excel(results, output_file, append=False):
    # Separate results into lists based on language
    html_css_results = []
    sql_results = []
    java_results = []
    java_maven_results = []  # Initialize this variable

    for result in results:
        if result['Language'] == 'HTML/CSS':
            html_css_results.append(
                {
                    'Name': result['Folder Name'],
                    'Cohort': '',  # Placeholder for Cohort
                    'HTML': '',  # Placeholder for HTML
                    'CSS': '',  # Placeholder for CSS
                    'Various Features': '',  # Placeholder for Various Features
                    'Multiple Pages': '',  # Placeholder for Multiple Pages
                    'JavaScript': '',  # Placeholder for JavaScript
                    'Different / Cool': '',  # Placeholder for Different / Cool
                    'Good Design': '',  # Placeholder for Good Design
                    'Git link': result['Repository'],  # Maps to Repository
                    'Comments': '',  # Placeholder for Comments
                    'Help provided': '',  # Placeholder for Help provided
                    'Total': '',  # Placeholder for Total
                    'Percentage': ''  # Placeholder for Percentage
                }
            )
        elif result['Language'] == 'SQL':
            sql_results.append(
                {
                    'Name': result['Folder Name'],
                    'Cohort': '',  # Placeholder for Cohort
                    'Schema': '',  # Placeholder for Schema
                    'Actual Database': '',  # Placeholder for Actual Database
                    'Data': '',  # Placeholder for Data
                    'Queries': '',  # Placeholder for Queries
                    'Git link': result['Repository'],  # Maps to Repository
                    'Comments': '',  # Placeholder for Comments
                    'Help provided': '',  # Placeholder for Help provided
                    'Total': '',  # Placeholder for Total
                    'Percentage': ''  # Placeholder for Percentage
                }
            )
        elif result['Language'] == 'Java':
            java_results.append(
                {
                    'Name': result['Folder Name'],
                    'Score': '',  # Placeholder for Score
                    'Percent': '',  # Placeholder for Percent
                    'Overall Comments': '',  # Placeholder for Overall Comments
                    'Git repo': result['Repository'],  # Maps to Repository
                    'Other comments': '',  # Placeholder for Other comments
                    'Produced a random move for the computer player': '',  # Placeholder
                    'Asks human for their choice': '',  # Placeholder
                    'Announces the winner': '',  # Placeholder
                    'Includes replay if it\'s a draw': '',  # Placeholder
                    'Uses objects': '',  # Placeholder
                    'Additional features / something cool': '',  # Placeholder
                    'Compiles and runs': '',  # Placeholder
                    'Syntax and formatting sound': '',  # Placeholder
                    'Implementation appropriate': '',  # Placeholder
                    'General OOP principles applied': '',  # Placeholder
                    'Inheritance / abstract classes / enums': '',  # Placeholder
                }
            )
        elif result['Language'] == 'Java-Maven':
            java_maven_results.append(  # Fix: change java_results to java_maven_results
                {
                    'Name': result['Folder Name'],
                    'Score': '',  # Placeholder for Score
                    'Percent': '',  # Placeholder for Percent
                    'Overall Comments': '',  # Placeholder for Overall Comments
                    'Git repo': result['Repository'],  # Maps to Repository
                    'Other comments': '',  # Placeholder for Other comments
                    'Compiles and runs': '',  # Placeholder
                    'Syntax and formatting sound': '',  # Placeholder
                    'Range of tests attempted (e.g. more than just auth)': '',  # Placeholder
                    'Page Object Pattern implemented': '',  # Placeholder
                    'Reusable methods in the Pages': '',  # Placeholder
                    'Tests well structured and include assertions': '',  # Placeholder
                    'Selectors good - not just copied from dev tools': '',  # Placeholder
                    'Good ideas / implementation for the harder tests (looping over items, data driven tests etc)': ''  # Placeholder
                }
            )

    # Convert results to DataFrames
    results_df =pd.DataFrame(results)
    HTML_df = pd.DataFrame(html_css_results)
    SQL_df = pd.DataFrame(sql_results)
    Java_df = pd.DataFrame(java_results)
    Java_Maven_df = pd.DataFrame(java_maven_results)

    # Mapping of languages to corresponding sheet names
    sheet_name_mapping = {
        'Results' : 'Results',
        'HTML/CSS': 'WebAssessment',
        'SQL': 'DatabaseAssessment',
        'Java': 'ProgrammingAssessment',
        'Java-Maven': 'AutomationAssessment'
    }

    if append and os.path.exists(output_file):
        # Open the workbook and append results to the respective sheet based on the language
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(output_file)
            
            # Append results for each language to the appropriate sheet
            for lang_df, sheet_name in zip([HTML_df, SQL_df, Java_df, Java_Maven_df], 
                                            sheet_name_mapping.values()):
                if not lang_df.empty:  # Only process if DataFrame is not empty
                    if sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        
                        # Check each row to determine the last non-empty row
                        startrow = 1  # Initialize startrow
                        for row in range(1, worksheet.max_row + 1):
                            if any(cell.value is not None for cell in worksheet[row]):
                                startrow = row + 1  # Move startrow to the next empty row

                        # Now append your DataFrame to the next empty row
                        lang_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow, header=False)
                    else:
                        # If the sheet doesn't exist, create a new one and write the DataFrame
                        lang_df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Create a new Excel file with results so more info can be found on results.
        with pd.ExcelWriter(os.path.join(os.getcwd(), 'repo_processing_results.xlsx'), engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name="Results")
    else:
        # If append is False, create a new Excel file with different sheets for each assessment category
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for lang_df, sheet_name in zip([results_df, HTML_df, SQL_df, Java_df, Java_Maven_df], 
                                            sheet_name_mapping.values()):
                if not lang_df.empty:  # Only process if DataFrame is not empty
                    lang_df.to_excel(writer, index=False, sheet_name=sheet_name)

