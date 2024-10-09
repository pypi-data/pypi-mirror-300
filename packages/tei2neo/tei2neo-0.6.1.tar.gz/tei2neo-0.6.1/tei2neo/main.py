import datetime
import os
import pathlib
import re
import sys
import itertools

import click
import git
from dotenv import load_dotenv
from py2neo import Graph, Node, Relationship
from semper_backend.utils import GraphUtils
from tei2neo import parse
from openpyxl import load_workbook

from .parse import import_categories, read_index
from .toc import TOC

load_dotenv()
VERBOSE = False


def xml_files_filter():
    def my_coll_filter(val):
        if "Register" in val:
            return None
        match = re.search(r"xml$", val, re.IGNORECASE)
        return match

    return my_coll_filter


def get_graph(ctx=None):
    if ctx is None:
        host = os.environ.get("NEO4J_HOST")
        port = os.environ.get("NEO4J_PORT")
        user = os.environ.get("NEO4J_USERNAME")
        password = os.environ.get("NEO4J_PASSWORD")
    else:
        host = ctx.obj["host"] or os.getenv("NEO4J_HOST")
        port = ctx.obj["port"] or os.getenv("NEO4J_PORT")
        user = ctx.obj["username"] or os.getenv("NEO4J_USERNAME")
        password = ctx.obj["password"] or os.getenv("NEO4J_PASSWORD")

    host = host or "localhost"
    port = port or 7687
    user = user or "neo4j"
    password = (
        password
        or os.environ.get("NEO4J_PASSWORD")
        or click.prompt(text=f"Password for {user}@{host}:{port}", hide_input=True)
    )

    graph = Graph(
        host=host,
        port=port,
        user=user,
        password=password,
    )
    return graph


def delete_toc(graph):
    """Deletes the Table Of Content"""
    ut = GraphUtils(graph)
    ut.delete_toc()


def import_file(filepath, graph, ctx, commit=None):
    ut = GraphUtils(graph)

    filepath = os.path.abspath(filepath)
    filename = os.path.basename(filepath)
    info = ut.get_info_for_filename(filename)
    if info and ctx.obj["skip-already-imported-files"]:
        print(f"Importing {filepath} - already imported")
        return

    ut.delete_graph_for_file(filename)

    # import
    if VERBOSE:
        print("parse START:", datetime.datetime.now())
    print(f"Importing {filepath}", end="", flush=True)
    try:
        doc, *_ = parse(filename=filepath, commit=commit)
    except Exception as exc:
        print(f" ERROR:{str(exc)}", flush=True)
        return
    tx = graph.begin()
    doc.save(tx)
    tx.commit()

    print(" ok", flush=True)
    # if VERBOSE: print("parse DONE:", datetime.datetime.now())

    # create the relationships within the document
    # ut.link_inner_relationships(filename)
    # if VERBOSE: print("link_inner_relationships DONE:", datetime.datetime.now())

    # connect the document to existing categories
    # ut.connect_to_categories(filename)
    # if VERBOSE: print("connect_to_categories DONE:", datetime.datetime.now())

    # create the unhyphened tokens
    paras = ut.paragraphs_for_filename(filename)
    for para in paras:
        if VERBOSE:
            print(para)
        tokens = ut.tokens_in_paragraph(para)
        ut.create_unhyphenated(tokens)
    if VERBOSE:
        print("create_unhyphenated DONE:", datetime.datetime.now())


def delete_all_categories(graph=None):
    if graph is None:
        graph = get_graph()

    ut = GraphUtils(graph)
    ut.delete_categories()


# default=os.environ.get('NEO4J_HOSTNAME')
@click.group()
@click.version_option()
@click.option(
    "-h",
    "--host",
    default=lambda: os.environ.get("NEO4J_HOSTNAME"),
    help="Hostname of your Neo4j database or NEO4J_HOSTNAME env",
)
@click.option(
    "-p",
    "--port",
    default=lambda: os.environ.get("NEO4J_PORT"),
    help="port of your Neo4j database. NEO4J_PORT env",
)
@click.option(
    "-u",
    "--username",
    default=lambda: os.environ.get("NEO4J_USERNAME"),
    show_default="neo4j",
    help="your Neo4j username. NEO4J_USERNAME env or neo4j",
)
@click.option(
    "--password",
    default=lambda: os.environ.get("NEO4J_PASSWORD"),
    help="password of your Neo4j database or NEO4J_PASSWORD env",
)
@click.option("-v", "--verbose", is_flag=True)
@click.pass_context
def cli(ctx, host=None, port=None, username=None, password=None, verbose=False):
    """This is the main entry point for the Semper CLI.
    It offers subcommands (defined below) to easily import TEI documents into the
    graph database.
    """

    if verbose:
        global VERBOSE
        VERBOSE = True

    ctx.ensure_object(dict)
    ctx.obj["host"] = host
    ctx.obj["port"] = port
    ctx.obj["username"] = username
    ctx.obj["password"] = password


def create_node(tx, element, parent_node=None):
    labels = element.get("labels")
    if not isinstance(labels, list):
        labels = [labels]

    attrs = {}
    for key, val in element.get("attrs").items():
        attrs[str(key)] = val

    node = Node(*labels, **attrs)
    tx.create(node)

    if parent_node:
        rel = Relationship(parent_node, "HAS", node)
        tx.create(rel)

    for relation in element.get("relations"):
        create_node(tx, relation, parent_node=node)


def import_categories_from_path(ctx, path=None):
    """Import the categores from a given directory"""
    if path is None:
        try:
            path = sys.argv[1]
        except IndexError:
            while True:
                path = input("please enter location for categories.xml: ")
                path = os.path.expanduser(path)
                if os.path.exists(path) and os.path.isfile(path):
                    break

    graph = get_graph(ctx)
    tx = graph.begin()
    import_categories(path, tx)
    tx.commit()


def get_repo_for_path(tei_path):
    tei_path = os.path.dirname(tei_path)
    try:
        repo = git.Repo(tei_path)
    except git.exc.InvalidGitRepositoryError:
        repo = git.Repo(tei_path, search_parent_directories=True)
    return repo


@cli.command("import-tei")
@click.option("-s", "--skip-existing", is_flag=True, help="skip already imported files")
@click.option(
    "-d",
    "--diff",
    help="Import all files of master that changed since HEAD~x or commit-hash",
)
@click.option(
    "-l",
    "--latest",
    is_flag=True,
    help="fetch from origin/master and import all files that changed in meantime.",
)
@click.option(
    "-c",
    "--commit",
    help="Import all files of a certain commit hash (back in time) compared to master",
)
@click.argument(
    "path",
    nargs=-1,
    type=click.Path(),
)
@click.pass_context
def import_tei(ctx, skip_existing, diff=None, latest=False, commit=None, path=None):
    """Import all xml files from a given directory.
    These files must fit the pattern and are ordered by their page number
    before they are imorted.
    """
    ctx.obj["skip-already-imported-files"] = skip_existing
    if not path:
        raise click.ClickException("Please provide a path or a filename")
    if not pathlib.Path(path[0]).exists:
        raise click.ClickException(f"Path {path[0]} does not exist")
    repo = get_repo_for_path(path[0])

    changed_files = []
    if latest:
        commit = "origin/HEAD"

    if commit is not None:
        # get latest commits from origin
        for remote in repo.remotes:
            remote.fetch()
        # checkout the version indicated
        repo.git.checkout(commit)
        for diff_obj in repo.commit().diff("master"):
            changed_files.append(os.path.join(repo.working_dir, diff_obj.a_path))

    if diff is not None:
        # get latest commits from origin
        for remote in repo.remotes:
            remote.fetch()
        # checkout the version indicated
        repo.git.checkout("master")
        for diff_obj in repo.commit().diff(diff):
            changed_files.append(os.path.join(repo.working_dir, diff_obj.a_path))

    if len(changed_files) == 0:
        # user provided number of individual files that should be updated
        if len(path) == 1 and os.path.isdir(path[0]):
            changed_files = get_files_for_path(path[0])
        else:
            changed_files = list(path)

    changed_files = list(filter(xml_files_filter(), changed_files))
    if len(changed_files) == 0:
        print("no changes in XML files detected.")
        return

    print(f"will update {len(changed_files)} file(s).")
    print(changed_files)

    graph = get_graph(ctx)

    repo.git.checkout("master")
    commit_obj = repo.commit()
    for filepath in changed_files:
        if VERBOSE:
            print("importing:", filepath, end="")
        import_file(filepath=filepath, graph=graph, ctx=ctx, commit=commit_obj)

    # checkout again master
    repo.git.checkout("master")
    # merge changes into master
    # repo.git.pull()


@cli.command("import-register")
@click.argument("path", type=click.Path(exists=True))
def import_index_from_path(path, graph=None):
    """Import all TEI-XML files containing register information from a given directory.
    The TEI files are parsed and then nodes and relations are created.
    """
    registers = [
        "orgregister.xml",
        "artefaktenregister.xml",
        "ortsregister.xml",
        "begriffsregister.xml",
        "personen.xml",
        "voelkerregister.xml",
        "biblioregister.xml",
    ]

    graph = get_graph()

    register_filepaths = []
    filepath = pathlib.Path(path)
    if filepath.is_file():
        register_filepaths.append(filepath)
    else:
        for register in registers:
            filepath = pathlib.Path(path) / register
            if filepath.exists():
                register_filepaths.append(filepath)
            else:
                print(f"*** not found: {filepath}")

    for register_filepath in register_filepaths:
        tx = graph.begin()
        print(f"reading {register_filepath}...", end="", flush=True)
        elements = read_index(str(register_filepath))
        for element in elements:
            create_node(tx, element)
        tx.commit()
        print(" ok")


@cli.command()
@click.pass_context
def create_indexes(ctx):
    """Executes all indexes in a cypher file, line by line"""

    thisdir = os.path.dirname(os.path.abspath(__file__))
    cypher_file = os.path.join(thisdir, "create_indexes.cypher")

    graph = get_graph(ctx)
    ut = GraphUtils(graph)
    new = 0
    existing = 0
    with open(cypher_file, "r", encoding="utf-8") as fh:
        for command in fh:
            try:
                ut.execute_cypher(command)
                print(command)
                new += 1
            except Exception as exc:
                if "already exists" in str(exc):
                    existing += 1
                    pass
                else:
                    raise ValueError(exc)
    click.echo(f"{new} indexes created, {existing} already exist.")


def get_files_for_path(path):
    filenames = []
    # get the last directory name as an indicator for the collection
    collection = os.path.basename(os.path.normpath(path))
    for root, _, files in os.walk(path):
        for filename in files:
            filepath = os.path.join(root, filename)
            filenames.append(filepath)

    return filenames


@cli.command()
@click.argument("cypher")
@click.pass_context
def execute_cypher(ctx, cypher):
    """Executes an arbitrary cyper command. Use at your own risk!"""

    graph = get_graph(ctx)

    ut = GraphUtils(graph)
    ut.execute_cypher(cypher)


@cli.command("import-toc")
@click.argument(
    "toc_filepath", type=click.Path(exists=True, readable=True, allow_dash=True)
)
def import_toc_from_path(toc_filepath, graph=None):
    """Import the table of content (TOC) from a given file"""

    toc = TOC.new_from_file(toc_filepath=toc_filepath)
    if graph is None:
        graph = get_graph()

    toc.save_toc_to_graph(graph=graph)


@cli.command("import-similar-paragraphs")
@click.argument(
    "para_filepath", type=click.Path(exists=True, readable=True, allow_dash=True)
)
def connect_similar_paragraphs(para_filepath, graph=None):
    """Connect similar paragraphs provided in an Excel sheet"""
    wb = load_workbook(para_filepath)
    sheet = wb[wb.sheetnames[0]]
    if graph is None:
        graph = get_graph()
    tx = graph.begin()

    for row in sheet.iter_rows(3):
        nodes = []
        for col in row:
            if col.value is None:
                break

            # come items contain " und ", that means the paragraphs are split over two pages
            items = col.value.split(" und ")
            for item in items:
                filename, xmlid = item.split("#")
                cypher = "MATCH (p :p) WHERE p.filename=$filename and p.`xml:id`=$xmlid RETURN p"
                cursor = tx.graph.run(
                    cypher, parameters={"filename": filename, "xmlid": xmlid}
                )
                node = cursor.evaluate()
                if node is None:
                    print(f"####### PARAGRAPH NOT FOUND: {filename}, {xmlid}")
                    continue
                nodes.append(node)
        for pair in itertools.combinations(nodes, 2):
            rel1 = Relationship(pair[0], "IS_SIMILAR_TO", pair[1])
            tx.create(rel1)
            rel2 = Relationship(pair[1], "IS_SIMILAR_TO", pair[0])
            tx.create(rel2)
    tx.commit()


if __name__ == "__main__":
    cli(obj={})
